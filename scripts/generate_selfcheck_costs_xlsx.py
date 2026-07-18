"""Generate an Excel workbook documenting the cost/leverage model behind the
TaVyro CEO / Executive Management Self-Check calculator.

Source of truth: public/tavyro-hr-health-check.html (updateCalc, calculateBoundGLHours,
flukCostPerPerson, compCostByRisk, scenario factors, question costs).

The workbook is interactive: change the input cells on the "Rechner" sheet and all
results (Brutto, Netto, ROI, Szenarien) recompute via formulas.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Constants from the calculator
# ---------------------------------------------------------------------------
WEEKS = 46
HOURLY_RATE = 420          # CHF / GL-Stunde
GL_CAP = 30               # max. entlastbare GL-Std/Woche (gesamt)
SCEN_CONS = 0.7
SCEN_REAL = 1.0
SCEN_AMBI = 1.3

# fluktuation cost per person by team size (thresholds)
FLUK_TIERS = [
    ("Team > 200", 28000),
    ("Team > 120 bis 200", 23000),
    ("Team > 50 bis 120", 19000),
    ("Team 5 bis 50", 15000),
]

# compliance base cost by risk level
COMP_BASE = {"low": 15000, "mid": 35000, "high": 70000}

# band multipliers (Prioritaetsband)
BAND_MULT = {
    "low":  {"ceo": 1.00, "fluk": 1.00, "comp": 1.00},
    "mid":  {"ceo": 1.04, "fluk": 1.06, "comp": 1.06},
    "high": {"ceo": 1.08, "fluk": 1.12, "comp": 1.12},
}

# 16 questions: id, phase, short label, full text, cost red, cost yellow, GL-h red, GL-h yellow
QUESTIONS = [
    (1, 1, "GL-Aufmerksamkeit fuer Personalthemen",
     "Wie stark binden Personalthemen Ihre persoenliche Aufmerksamkeit, obwohl sie auf Fuehrungs- oder Organisationsebene geloest werden sollten?", 118000, 40500, 90, 35),
    (2, 1, "Entscheidungsreife HR-Empfehlungen",
     "Erhalten Sie bei People- und Organisationsthemen entscheidungsreife Empfehlungen oder primaer operative Einzelthemen?", 63000, 18000, 45, 18),
    (3, 1, "Rollen-/Prozessklarheit & Skalierbarkeit",
     "Sind Rollen, Verantwortungen und Prozesse so klar definiert, dass die Organisation ohne wesentliche Reibungsverluste skalieren kann?", 78500, 22500, 35, 12),
    (4, 1, "Eigenstaendige Kaderentscheide",
     "Trifft Ihr Kader Personalentscheide innerhalb klarer Leitplanken eigenstaendig oder werden heikle Themen regelmaessig an Sie eskaliert?", 52500, 15000, 60, 22),
    (5, 1, "People-/Organisationsplanung (3 Jahre)",
     "Haben Sie eine klare People und Organisationsplanung, die Ihre Geschaeftsziele der naechsten drei Jahre unterstuetzt?", 87500, 22500, 30, 10),
    (6, 2, "Systematischer Kompetenzaufbau (inkl. AI)",
     "Werden kritische Faehigkeiten im Unternehmen systematisch aufgebaut - auch in digitalen und AI-nahen Kompetenzfeldern?", 70000, 18000, 15, 5),
    (7, 2, "Rahmen fuer Konflikte/Spannungen",
     "Gibt es in Ihrer Organisation einen vertrauenswuerdigen Rahmen, um Konflikte, Spannungen oder Grenzthemen fruehzeitig anzusprechen?", 61000, 15000, 20, 7),
    (8, 2, "Nachfolge Schluesselpositionen",
     "Ist die Nachfolge fuer Schluesselpositionen so geregelt, dass ein Abgang keine wesentliche Betriebsstoerung verursacht?", 140000, 30000, 25, 8),
    (9, 2, "Arbeitgeber-Positionierung",
     "Ist Ihre Positionierung als Arbeitgeber so differenziert, dass sich qualifizierte Talente aktiv bei Ihnen bewerben?", 52500, 15000, 10, 4),
    (10, 2, "Strukturierte Rekrutierung",
     "Sind Ihre Rekrutierungsprozesse so strukturiert, dass qualifizierte Kandidaten nicht durch Verzoegerungen verloren gehen?", 43500, 12000, 15, 5),
    (11, 2, "Verguetungsstrukturen",
     "Sind Ihre Verguetungsstrukturen marktgerecht, transparent und nachvollziehbar fuer das Team?", 70000, 15000, 12, 4),
    (12, 2, "Onboarding/Integration",
     "Sind neue Mitarbeitende nach 4 Wochen so integriert, dass sie vollstaendig produktiv und engagiert in ihrer Rolle arbeiten?", 52500, 12000, 18, 6),
    (13, 3, "Digitale HR-Daten/Dokumente",
     "Sind relevante Personal-Kennzahlen und Dokumente digital verfuegbar, ohne Aufwand in verteilten Ablagesystemen zu suchen?", 35000, 9000, 10, 3),
    (14, 3, "Absenzmanagement",
     "Werden krankheitsbedingte Absenzen aktiv begleitet, sodass Ausfallzeiten und Versicherungskosten kontrolliert bleiben?", 61000, 15000, 8, 3),
    (15, 3, "Vertraege/Reglemente/Datenschutz",
     "Sind Vertraege, Reglemente, Datenschutz und arbeitsrechtliche Grundlagen aktuell und konsistent umgesetzt?", 35000, 7500, 10, 3),
    (16, 3, "Arbeitszeiterfassung",
     "Ist Ihre Arbeitszeiterfassung so aufgesetzt, dass Sie Pflichten und Nachweise jederzeit belastbar erbringen koennen?", 35000, 7500, 8, 2),
]

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------
BRAND = "2D6A4F"
BRAND_LIGHT = "E1F5EE"
GREY = "F2F2F0"
DARK = "1A1A18"

f_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
f_h2 = Font(name="Calibri", size=12, bold=True, color=DARK)
f_hdr = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
f_bold = Font(name="Calibri", size=10, bold=True, color=DARK)
f_norm = Font(name="Calibri", size=10, color=DARK)
f_input = Font(name="Calibri", size=10, bold=True, color="1F4E79")
f_note = Font(name="Calibri", size=9, italic=True, color="666666")

fill_brand = PatternFill("solid", fgColor=BRAND)
fill_hdr = PatternFill("solid", fgColor=BRAND)
fill_light = PatternFill("solid", fgColor=BRAND_LIGHT)
fill_grey = PatternFill("solid", fgColor=GREY)
fill_input = PatternFill("solid", fgColor="FFF6D6")

thin = Side(style="thin", color="D9D9D9")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center")
right = Alignment(horizontal="right", vertical="center")

CHF_FMT = '#,##0\\ "CHF";-#,##0\\ "CHF"'
CHF_FMT2 = '"CHF"\\ #,##0'


def style_header_row(ws, row, ncols, start=1):
    for c in range(start, start + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = f_hdr
        cell.fill = fill_hdr
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = border_all


def title_bar(ws, text, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=text)
    c.font = f_title
    c.fill = fill_brand
    c.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 30


wb = Workbook()

# ===========================================================================
# Sheet 1: Rechner (interactive)
# ===========================================================================
ws = wb.active
ws.title = "Rechner"
title_bar(ws, "TaVyro Self-Check - Orientierungsrechnung (interaktiv)", 4)
ws.column_dimensions["A"].width = 46
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 4
ws.column_dimensions["D"].width = 60

r = 3
ws.cell(row=r, column=1, value="EINGABEN (Felder gelb - aenderbar)").font = f_h2
r += 1
inputs = [
    ("Team-Groesse (Mitarbeitende)", 100, "5 - 280"),
    ("Betroffene GL-Mitglieder", 2, "1 - 6"),
    ("Entlastbare GL-Zeit / Mitglied / Woche (h)", 6, "0 - 8"),
    ("Vermiedene Abgaenge (Personen / Jahr)", 2, "0 - 10"),
    ("TaVyro-Investition / Jahr (CHF)", 80000, "12'000 - 180'000"),
    ("Risikoniveau (Governance/Compliance)", "mid", "low / mid / high"),
    ("Prioritaetsband (aus Self-Check)", "mid", "low / mid / high"),
]
input_rows = {}
for label, val, hint in inputs:
    ws.cell(row=r, column=1, value=label).font = f_norm
    cell = ws.cell(row=r, column=2, value=val)
    cell.font = f_input
    cell.fill = fill_input
    cell.border = border_all
    cell.alignment = center
    if isinstance(val, int) and "CHF" in label:
        cell.number_format = CHF_FMT2
    ws.cell(row=r, column=4, value="zulaessig: " + hint).font = f_note
    input_rows[label] = r
    r += 1

TEAM = f"B{input_rows['Team-Groesse (Mitarbeitende)']}"
GLM = f"B{input_rows['Betroffene GL-Mitglieder']}"
GLH = f"B{input_rows['Entlastbare GL-Zeit / Mitglied / Woche (h)']}"
FLUK = f"B{input_rows['Vermiedene Abgaenge (Personen / Jahr)']}"
INVEST = f"B{input_rows['TaVyro-Investition / Jahr (CHF)']}"
RISK = f"B{input_rows['Risikoniveau (Governance/Compliance)']}"
BAND = f"B{input_rows['Prioritaetsband (aus Self-Check)']}"

# dropdowns
dv = DataValidation(type="list", formula1='"low,mid,high"', allow_blank=False)
ws.add_data_validation(dv)
dv.add(ws[RISK])
dv.add(ws[BAND])

r += 1
ws.cell(row=r, column=1, value="ZWISCHENGROESSEN").font = f_h2
r += 1


def calc_row(label, formula, fmt=None, bold=False, fill=None):
    global r
    ws.cell(row=r, column=1, value=label).font = f_bold if bold else f_norm
    cell = ws.cell(row=r, column=2, value=formula)
    cell.font = f_bold if bold else f_norm
    cell.alignment = right
    cell.border = border_all
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
        ws.cell(row=r, column=1).fill = fill
    ref = f"B{r}"
    r += 1
    return ref


GLTOT = calc_row("Entlastbare GL-Std / Woche (max. 30)", f"=MIN({GL_CAP},{GLM}*{GLH})")
BCEO = calc_row("Band-Faktor CEO/GL", f'=IF({BAND}="high",1.08,IF({BAND}="mid",1.04,1))')
BFLK = calc_row("Band-Faktor Fluktuation", f'=IF({BAND}="high",1.12,IF({BAND}="mid",1.06,1))')
BCMP = calc_row("Band-Faktor Compliance", f'=IF({BAND}="high",1.12,IF({BAND}="mid",1.06,1))')
CPP = calc_row("Fluktuationskosten / Person", f"=IF({TEAM}>200,28000,IF({TEAM}>120,23000,IF({TEAM}>50,19000,15000)))", CHF_FMT2)
BASEC = calc_row("Compliance-Basiswert (Risiko)", f'=IF({RISK}="low",15000,IF({RISK}="high",70000,35000))', CHF_FMT2)
NORM = calc_row("Normierung Teamgroesse (0-1)", f"=MAX(0,MIN(1,({TEAM}-5)/275))", "0.000")
CMULT = calc_row("Compliance-Multiplikator (1 + norm*0.8)", f"=1+{NORM}*0.8", "0.000")

r += 1
ws.cell(row=r, column=1, value="HEBEL / EINSPARUNG (Brutto)").font = f_h2
r += 1
CEOS = calc_row("Entlastung CEO/GL", f"=ROUND({GLTOT}*{WEEKS}*{HOURLY_RATE}*{BCEO},0)", CHF_FMT2)
FLKS = calc_row("Vermiedene Fluktuation", f"=ROUND({FLUK}*{CPP}*{BFLK},0)", CHF_FMT2)
CMPS = calc_row("Risiko-/Compliance-Minimierung", f"=ROUND({BASEC}*{CMULT}*{BCMP},0)", CHF_FMT2)
TOTP = calc_row("Brutto-Potenzial / Jahr", f"={CEOS}+{FLKS}+{CMPS}", CHF_FMT2, bold=True, fill=fill_light)

r += 1
ws.cell(row=r, column=1, value="ERGEBNIS").font = f_h2
r += 1
calc_row("abzgl. TaVyro-Investition", f"=-{INVEST}", CHF_FMT2)
NETTO = calc_row("Netto-Effekt / Jahr (Realistisch)", f"={TOTP}-{INVEST}", CHF_FMT2, bold=True, fill=fill_light)
calc_row("Verhaeltnis Potenzial / Investition", f'=IF({INVEST}>0,{TOTP}/{INVEST},"-")', "0.0\\x")

r += 1
ws.cell(row=r, column=1, value="SZENARIEN (Netto-Effekt)").font = f_h2
r += 1
calc_row("Konservativ (Brutto x 0.7 - Investition)", f"=ROUND({TOTP}*{SCEN_CONS},0)-{INVEST}", CHF_FMT2)
calc_row("Realistisch (Brutto x 1.0 - Investition)", f"={NETTO}", CHF_FMT2, bold=True)
calc_row("Ambitioniert (Brutto x 1.3 - Investition)", f"=ROUND({TOTP}*{SCEN_AMBI},0)-{INVEST}", CHF_FMT2)

r += 1
note = ("Modell: Entlastung CEO/GL = GL-Std/Woche (max. 30) x 46 Wochen x CHF 420/Std x Band-Faktor. "
        "Vermiedene Fluktuation = Abgaenge x Kosten/Person (teamgroessenabhaengig) x Band-Faktor. "
        "Risiko-/Compliance = Basiswert (Risiko) x (1 + Norm*0.8) x Band-Faktor. "
        "Qualitative Prioritaet und monetaere Rechnung sind getrennte Modelle; bei hohem Prioritaetsindex "
        "werden Hebelannahmen moderat angehoben, bleiben jedoch konservativ. Werte sind Orientierungsgroessen.")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
nc = ws.cell(row=r, column=1, value=note)
nc.font = f_note
nc.alignment = wrap
ws.row_dimensions[r].height = 58

# ===========================================================================
# Sheet 2: Optionen-Matrix (Band x Risiko, bei aktuellen Rechner-Eingaben)
# ===========================================================================
ws2 = wb.create_sheet("Optionen-Matrix")
title_bar(ws2, "Alle Optionen: Prioritaetsband x Risikoniveau (Eingaben aus 'Rechner')", 10)
headers = ["Prioritaetsband", "Risikoniveau", "Entlastung CEO/GL", "Vermiedene Fluktuation",
           "Risiko/Compliance", "Brutto-Potenzial", "Netto (Real.)", "Konservativ", "Realistisch", "Ambitioniert"]
for i, h in enumerate(headers, start=1):
    ws2.cell(row=3, column=i, value=h)
style_header_row(ws2, 3, len(headers))
ws2.column_dimensions["A"].width = 16
ws2.column_dimensions["B"].width = 14
for col in "CDEFGHIJ":
    ws2.column_dimensions[col].width = 17

row = 4
for band in ["low", "mid", "high"]:
    bceo = BAND_MULT[band]["ceo"]
    bflk = BAND_MULT[band]["fluk"]
    bcmp = BAND_MULT[band]["comp"]
    for risk in ["low", "mid", "high"]:
        base = COMP_BASE[risk]
        gltot = f"MIN({GL_CAP},Rechner!{GLM}*Rechner!{GLH})"
        cpp = f"IF(Rechner!{TEAM}>200,28000,IF(Rechner!{TEAM}>120,23000,IF(Rechner!{TEAM}>50,19000,15000)))"
        cmult = f"(1+MAX(0,MIN(1,(Rechner!{TEAM}-5)/275))*0.8)"
        ceos = f"ROUND({gltot}*{WEEKS}*{HOURLY_RATE}*{bceo},0)"
        flks = f"ROUND(Rechner!{FLUK}*{cpp}*{bflk},0)"
        cmps = f"ROUND({base}*{cmult}*{bcmp},0)"
        tot = f"({ceos}+{flks}+{cmps})"
        ws2.cell(row=row, column=1, value=band).alignment = center
        ws2.cell(row=row, column=2, value=risk).alignment = center
        ws2.cell(row=row, column=3, value=f"={ceos}")
        ws2.cell(row=row, column=4, value=f"={flks}")
        ws2.cell(row=row, column=5, value=f"={cmps}")
        ws2.cell(row=row, column=6, value=f"={tot}")
        ws2.cell(row=row, column=7, value=f"={tot}-Rechner!{INVEST}")
        ws2.cell(row=row, column=8, value=f"=ROUND({tot}*{SCEN_CONS},0)-Rechner!{INVEST}")
        ws2.cell(row=row, column=9, value=f"={tot}-Rechner!{INVEST}")
        ws2.cell(row=row, column=10, value=f"=ROUND({tot}*{SCEN_AMBI},0)-Rechner!{INVEST}")
        for col in range(3, 11):
            cell = ws2.cell(row=row, column=col)
            cell.number_format = CHF_FMT2
            cell.alignment = right
            cell.border = border_all
        for col in range(1, 3):
            ws2.cell(row=row, column=col).border = border_all
        if band == "mid" and risk == "mid":
            for col in range(1, 11):
                ws2.cell(row=row, column=col).fill = fill_light
        row += 1

ws2.cell(row=row + 1, column=1,
         value="Hinweis: Diese Matrix nutzt die aktuellen Werte fuer Team, GL-Mitglieder, GL-Stunden, "
               "Abgaenge und Investition aus dem Tabellenblatt 'Rechner'. Hervorgehoben: mid/mid (Basisszenario).").font = f_note
ws2.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=10)

# ===========================================================================
# Sheet 3: Fragen & Hebel (16 Fragen)
# ===========================================================================
ws3 = wb.create_sheet("Fragen & Hebel")
title_bar(ws3, "16 Self-Check-Fragen - Hebelwerte & gebundene GL-Stunden", 9)
headers3 = ["Nr", "Phase", "Thema", "Frage (Volltext)",
            "Hebel KRITISCH (rot)", "Hebel TEILWEISE (gelb)",
            "Hebel gerundet rot (5'000)", "GL-Std/J rot", "GL-Std/J gelb"]
for i, h in enumerate(headers3, start=1):
    ws3.cell(row=3, column=i, value=h)
style_header_row(ws3, 3, len(headers3))
widths3 = [5, 7, 34, 70, 16, 16, 16, 12, 12]
for i, w in enumerate(widths3, start=1):
    ws3.column_dimensions[get_column_letter(i)].width = w

row = 4
phase_names = {1: "Phase 1", 2: "Phase 2", 3: "Phase 3"}
for (qid, phase, label, text, cr, cy, hr, hy) in QUESTIONS:
    ws3.cell(row=row, column=1, value=qid).alignment = center
    ws3.cell(row=row, column=2, value=phase_names[phase]).alignment = center
    ws3.cell(row=row, column=3, value=label).alignment = wrap
    ws3.cell(row=row, column=4, value=text).alignment = wrap
    ws3.cell(row=row, column=5, value=cr).number_format = CHF_FMT2
    ws3.cell(row=row, column=6, value=cy).number_format = CHF_FMT2
    ws3.cell(row=row, column=7, value=round(cr / 5000) * 5000).number_format = CHF_FMT2
    ws3.cell(row=row, column=8, value=hr).alignment = center
    ws3.cell(row=row, column=9, value=hy).alignment = center
    for col in range(1, 10):
        ws3.cell(row=row, column=col).border = border_all
        ws3.cell(row=row, column=col).font = f_norm
    ws3.row_dimensions[row].height = 42
    row += 1

# totals
ws3.cell(row=row, column=4, value="Summe (theoretisch, NICHT summierbar mit Rechner)").font = f_bold
ws3.cell(row=row, column=5, value=f"=SUM(E4:E{row-1})").number_format = CHF_FMT2
ws3.cell(row=row, column=6, value=f"=SUM(F4:F{row-1})").number_format = CHF_FMT2
ws3.cell(row=row, column=8, value=f"=SUM(H4:H{row-1})").alignment = center
ws3.cell(row=row, column=9, value=f"=SUM(I4:I{row-1})").alignment = center
for col in range(4, 10):
    ws3.cell(row=row, column=col).fill = fill_grey
    ws3.cell(row=row, column=col).border = border_all
row += 2
ws3.cell(row=row, column=1,
         value="Hinweis: Hebelwerte der einzelnen Fragen sind themenspezifische Richtwerte, NICHT summierbar "
               "und nicht 1:1 mit der monetaeren Orientierungsrechnung gleichzusetzen. Die GL-Stunden je Frage "
               "speisen die Dashboard-Kennzahl 'gebundene GL-Zeit' (Std x CHF 420).").font = f_note
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)

# ===========================================================================
# Sheet 4: Annahmen & Konstanten
# ===========================================================================
ws4 = wb.create_sheet("Annahmen & Konstanten")
title_bar(ws4, "Modellannahmen & Konstanten", 3)
ws4.column_dimensions["A"].width = 44
ws4.column_dimensions["B"].width = 20
ws4.column_dimensions["C"].width = 40

row = 3


def kv(label, value, note="", fmt=None):
    global row
    ws4.cell(row=row, column=1, value=label).font = f_norm
    c = ws4.cell(row=row, column=2, value=value)
    c.font = f_bold
    c.alignment = right
    if fmt:
        c.number_format = fmt
    ws4.cell(row=row, column=3, value=note).font = f_note
    for col in range(1, 4):
        ws4.cell(row=row, column=col).border = border_all
    row += 1


def section(t):
    global row
    ws4.cell(row=row, column=1, value=t).font = f_h2
    row += 1


section("Allgemein")
kv("Arbeitswochen pro Jahr", WEEKS)
kv("GL-Stundensatz", HOURLY_RATE, "CHF / Stunde", CHF_FMT2)
kv("Max. entlastbare GL-Std / Woche (Cap)", GL_CAP, "h/Woche gesamt")

row += 1
section("Fluktuationskosten pro Person (nach Teamgroesse)")
for label, val in FLUK_TIERS:
    kv(label, val, "CHF / Person", CHF_FMT2)

row += 1
section("Compliance-Basiswert (nach Risikoniveau)")
for risk, val in COMP_BASE.items():
    kv(f"Risiko = {risk}", val, "CHF", CHF_FMT2)
kv("Compliance-Multiplikator", "1 + Norm * 0.8", "Norm = (Team-5)/275, gekappt 0-1")

row += 1
section("Band-Faktoren (Prioritaetsband aus Self-Check)")
ws4.cell(row=row, column=1, value="Band").font = f_hdr
ws4.cell(row=row, column=1).fill = fill_hdr
ws4.cell(row=row, column=2, value="CEO/GL").font = f_hdr
ws4.cell(row=row, column=2).fill = fill_hdr
ws4.cell(row=row, column=3, value="Fluktuation / Compliance").font = f_hdr
ws4.cell(row=row, column=3).fill = fill_hdr
for col in range(1, 4):
    ws4.cell(row=row, column=col).alignment = center
    ws4.cell(row=row, column=col).border = border_all
row += 1
for band, m in BAND_MULT.items():
    ws4.cell(row=row, column=1, value=band).font = f_norm
    ws4.cell(row=row, column=2, value=m["ceo"]).font = f_norm
    ws4.cell(row=row, column=3, value=f'{m["fluk"]} / {m["comp"]}').font = f_norm
    for col in range(1, 4):
        ws4.cell(row=row, column=col).border = border_all
        ws4.cell(row=row, column=col).alignment = center
    row += 1

row += 1
section("Szenario-Faktoren (auf Brutto-Potenzial)")
kv("Konservativ", SCEN_CONS, "x Brutto, dann minus Investition")
kv("Realistisch", SCEN_REAL, "x Brutto, dann minus Investition")
kv("Ambitioniert", SCEN_AMBI, "x Brutto, dann minus Investition")

row += 2
ws4.cell(row=row, column=1,
         value="Quelle: public/tavyro-hr-health-check.html (Stand der Berechnungslogik). "
               "Erstellt als interne Arbeitsgrundlage.").font = f_note
ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

# freeze panes
ws.freeze_panes = "A3"
ws2.freeze_panes = "A4"
ws3.freeze_panes = "A4"

# ---------------------------------------------------------------------------
out_dir = os.path.expanduser(
    "~/Library/CloudStorage/OneDrive-Persönlich/TaVyro GmbH/CEO_GL_Selfcheck"
)
if not os.path.isdir(out_dir):
    # Fallback: write next to the repo if the OneDrive folder is unavailable
    out_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "CEO_GL_Selfcheck")
    os.makedirs(out_dir, exist_ok=True)
out_path = os.path.join(out_dir, "TaVyro_Selfcheck_Kostenkalkulation.xlsx")
wb.save(out_path)
print("Saved:", out_path)
